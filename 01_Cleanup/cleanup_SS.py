import openpyxl
import os
from pathlib import Path
import re
from datetime import datetime

# Directories
INPUT_EXCEL = "master.xlsx"
RAW_TEXT_DIR = "txt_files_raw"
OUTPUT_DIR = "txt_files_cleaned_SS"
OUTPUT_EXCEL = "cleanup_SS_log.xlsx"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load the input Excel file (DOCUMENTS tab) - data_only=True to read calculated values, not formulas
wb_input = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
ws_input = wb_input["DOCUMENTS"]

# Create a new Excel file for tracking warnings and word counts
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.title = "Cleanup_Log"

# Set up headers in output Excel
headers = ["Doc_ID", "Firm", "Title", "Status", "Warnings", "Cleaned_Word_Count"]
for col, header in enumerate(headers, start=1):
    ws_output.cell(row=1, column=col, value=header)

# Find column indices in input Excel
firm_col = None
title_col = None
doc_id_col = None

header_row = ws_input[1]
for cell in header_row:
    if cell.value:
        if "Firm" in str(cell.value):
            firm_col = cell.column
        elif "Title" in str(cell.value) or "title" in str(cell.value):
            title_col = cell.column
        elif "Doc_ID" in str(cell.value) or "doc_id" in str(cell.value):
            doc_id_col = cell.column

if not firm_col or not title_col or not doc_id_col:
    print("Error: Could not find required columns (Firm, Title, Doc_ID) in DOCUMENTS tab")
    exit(1)

print(f"Output directory: {OUTPUT_DIR}")
print(f"Output Excel: {OUTPUT_EXCEL}")
print(f"Firm col: {openpyxl.utils.get_column_letter(firm_col)}, Title col: {openpyxl.utils.get_column_letter(title_col)}, Doc_ID col: {openpyxl.utils.get_column_letter(doc_id_col)}")
print("-" * 80)

# Process rows
processed = 0
skipped = 0
failed = 0
output_row = 2

for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=False), start=2):
    try:
        # Get firm, title, and doc_id
        firm_cell = ws_input.cell(row=row_idx, column=firm_col)
        title_cell = ws_input.cell(row=row_idx, column=title_col)
        doc_id_cell = ws_input.cell(row=row_idx, column=doc_id_col)
        
        firm = firm_cell.value
        title = title_cell.value
        doc_id = doc_id_cell.value
        
        # Skip if not SS
        if firm != "SS":
            skipped += 1
            continue
        
        # Skip if missing title or doc_id
        if not title or not doc_id:
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value="Missing title or doc_id")
            output_row += 1
            failed += 1
            continue
        
        # Construct raw text file path (Doc_ID already includes firm prefix like "SS_01")
        raw_file_path = os.path.join(RAW_TEXT_DIR, f"{doc_id}.txt")
        
        # Check if file exists
        if not os.path.exists(raw_file_path):
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value=f"Raw text file not found: {raw_file_path}")
            output_row += 1
            failed += 1
            continue
        
        # Read raw text
        with open(raw_file_path, 'r', encoding='utf-8') as f:
            raw_text = f.read()
        
        # ===== SS-SPECIFIC CLEANUP =====
        warnings = []
        
        # Find start point: 
        # Look for "SectionsIn this article" OR "| [number] min read" (whichever comes LATER)
        # If neither found, skip lines: "Skip to Main Content", "View All Results", "Search"
        
        sections_pattern = r'^SectionsIn this article$'
        min_read_pattern = r'^\| \d+ mins? read$'
        
        sections_match = re.search(sections_pattern, raw_text, re.MULTILINE)
        min_read_match = re.search(min_read_pattern, raw_text, re.MULTILINE)
        
        start_pos = 0
        start_marker_found = False
        
        if sections_match and min_read_match:
            # Use whichever comes LATER
            if sections_match.start() > min_read_match.start():
                start_pos = sections_match.end()
                start_marker_found = True
            else:
                start_pos = min_read_match.end()
                start_marker_found = True
        elif sections_match:
            start_pos = sections_match.end()
            start_marker_found = True
        elif min_read_match:
            start_pos = min_read_match.end()
            start_marker_found = True
        
        # If main markers not found, skip the header lines
        if not start_marker_found:
            # Remove lines: "Skip to Main Content", "View All Results", "Search"
            lines = raw_text.split('\n')
            skip_lines = {"skip to main content", "view all results", "search"}
            
            skip_until_idx = 0
            for idx, line in enumerate(lines):
                if line.strip().lower() not in skip_lines:
                    skip_until_idx = idx
                    break
            
            start_pos = 0
            raw_text = '\n'.join(lines[skip_until_idx:])
            warnings.append("Main start markers not found - skipped header lines")
        
        # Skip to next line after start marker
        if start_marker_found:
            next_line_start = raw_text.find('\n', start_pos)
            if next_line_start != -1:
                start_pos = next_line_start + 1
            else:
                start_pos = len(raw_text)
            
            # Skip leading whitespace
            while start_pos < len(raw_text) and raw_text[start_pos] in ' \n\t':
                start_pos += 1
        
        body_text_from_start = raw_text[start_pos:]
        
        # Find end point: "Related Insights", "Methodology", or "About Spencer Stuart" (whichever comes first)
        # Also find "Author" / "Authors" but use the LAST occurrence
        # Case-insensitive, but must be alone on one line
        related_insights_pattern = r'^Related Insights$'
        methodology_pattern = r'^Methodology$'
        about_spencer_pattern = r'^About Spencer Stuart$'
        author_pattern = r'^Authors?$'  # Matches both "Author" and "Authors"
        
        related_insights_match = re.search(related_insights_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        methodology_match = re.search(methodology_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        about_spencer_match = re.search(about_spencer_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        
        # For Author/Authors, find ALL occurrences and use the LAST one
        author_matches = list(re.finditer(author_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE))
        author_match = author_matches[-1] if author_matches else None
        
        end_pos = len(body_text_from_start)
        
        # Collect all matches and find the earliest one
        matches = []
        if related_insights_match:
            matches.append(("Related Insights", related_insights_match.start()))
        if methodology_match:
            matches.append(("Methodology", methodology_match.start()))
        if about_spencer_match:
            matches.append(("About Spencer Stuart", about_spencer_match.start()))
        if author_match:
            matches.append(("Author(s) [LAST]", author_match.start()))
        
        if matches:
            # Sort by position and use the first one
            matches.sort(key=lambda x: x[1])
            end_pos = matches[0][1]
        else:
            warnings.append("No end marker found (Related Insights/Methodology/About Spencer Stuart/Author(s))")
        
        # Extract body text
        body_text = body_text_from_start[:end_pos].strip()
        
        # Construct cleaned text: Title + body
        cleaned_text = f"{title}\n\n{body_text}"
        
        # Count words (all words in cleaned text)
        word_count = len(cleaned_text.split())
        
        # Write cleaned text to output file
        output_file_path = os.path.join(OUTPUT_DIR, f"{doc_id}.txt")
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned_text)
        
        # Write to output Excel
        ws_output.cell(row=output_row, column=1, value=doc_id)
        ws_output.cell(row=output_row, column=2, value=firm)
        ws_output.cell(row=output_row, column=3, value=title)
        ws_output.cell(row=output_row, column=4, value="SUCCESS")
        ws_output.cell(row=output_row, column=5, value=" | ".join(warnings) if warnings else "")
        ws_output.cell(row=output_row, column=6, value=word_count)
        output_row += 1
        
        processed += 1
        status_str = f"✓ Processed | Words: {word_count}"
        if warnings:
            status_str += f" | Warnings: {'; '.join(warnings)}"
        print(f"Row {row_idx} (Doc_ID: {doc_id}): {status_str}")
    
    except Exception as e:
        ws_output.cell(row=output_row, column=1, value=doc_id if 'doc_id' in locals() else "")
        ws_output.cell(row=output_row, column=2, value=firm if 'firm' in locals() else "")
        ws_output.cell(row=output_row, column=3, value=title if 'title' in locals() else "")
        ws_output.cell(row=output_row, column=4, value="FAILED")
        ws_output.cell(row=output_row, column=5, value=f"Error during cleanup: {str(e)}")
        output_row += 1
        print(f"Row {row_idx}: Error during cleanup - {str(e)}")
        failed += 1
        continue

# Save the output Excel
wb_output.save(OUTPUT_EXCEL)

print("-" * 80)
print(f"Cleanup complete!")
print(f"Processed: {processed}")
print(f"Skipped (non-SS): {skipped}")
print(f"Failed: {failed}")
print(f"Cleaned files saved to: {OUTPUT_DIR}/")
print(f"Log saved to: {OUTPUT_EXCEL}")

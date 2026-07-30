import openpyxl
import os
from pathlib import Path
import re
from datetime import datetime

# Directories
INPUT_EXCEL = "master.xlsx"
RAW_TEXT_DIR = "txt_files_raw"
OUTPUT_DIR = "txt_files_cleaned_RR"
OUTPUT_EXCEL = "cleanup_RR_log.xlsx"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load the input Excel file (DOCUMENTS tab) - data_only=True to read calculated values, not formulas
wb_input = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
ws_input = wb_input["DOCUMENTS"]

# Create a new Excel file for tracking warnings and word counts
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.title = "Cleanup_Log"

# Set up headers in output Excel
headers = ["Doc_ID", "Firm", "Title", "Status", "Warnings", "Cleaned_Word_Count"]
for col, header in enumerate(headers, start=1):
    ws_output.cell(row=1, column=col, value=header)

# Find column indices in input Excel
firm_col = None
title_col = None
doc_id_col = None

header_row = ws_input[1]
for cell in header_row:
    if cell.value:
        if "Firm" in str(cell.value):
            firm_col = cell.column
        elif "Title" in str(cell.value) or "title" in str(cell.value):
            title_col = cell.column
        elif "Doc_ID" in str(cell.value) or "doc_id" in str(cell.value):
            doc_id_col = cell.column

if not firm_col or not title_col or not doc_id_col:
    print("Error: Could not find required columns (Firm, Title, Doc_ID) in DOCUMENTS tab")
    exit(1)

print(f"Output directory: {OUTPUT_DIR}")
print(f"Output Excel: {OUTPUT_EXCEL}")
print(f"Firm col: {openpyxl.utils.get_column_letter(firm_col)}, Title col: {openpyxl.utils.get_column_letter(title_col)}, Doc_ID col: {openpyxl.utils.get_column_letter(doc_id_col)}")
print("-" * 80)

# Process rows
processed = 0
skipped = 0
failed = 0
output_row = 2

for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=False), start=2):
    try:
        # Get firm, title, and doc_id
        firm_cell = ws_input.cell(row=row_idx, column=firm_col)
        title_cell = ws_input.cell(row=row_idx, column=title_col)
        doc_id_cell = ws_input.cell(row=row_idx, column=doc_id_col)
        
        firm = firm_cell.value
        title = title_cell.value
        doc_id = doc_id_cell.value
        
        # Skip if not RR
        if firm != "RR":
            skipped += 1
            continue
        
        # Skip if missing title or doc_id
        if not title or not doc_id:
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value="Missing title or doc_id")
            output_row += 1
            failed += 1
            continue
        
        # Construct raw text file path (Doc_ID already includes firm prefix like "RR_01")
        raw_file_path = os.path.join(RAW_TEXT_DIR, f"{doc_id}.txt")
        
        # Check if file exists
        if not os.path.exists(raw_file_path):
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value=f"Raw text file not found: {raw_file_path}")
            output_row += 1
            failed += 1
            continue
        
        # Read raw text
        with open(raw_file_path, 'r', encoding='utf-8') as f:
            raw_text = f.read()
        
        # ===== RR-SPECIFIC CLEANUP =====
        warnings = []
        
        # Find start point: 
        # 1. Skip first line
        # 2. Search for title repetition (with nothing after it - end of line)
        # 3. Start from line after the title repetition
        
        lines = raw_text.split('\n')
        
        if len(lines) > 1:
            # Skip first line and search for title in remaining text
            remaining_text = '\n'.join(lines[1:])
            
            # Search for the title at the end of a line (with nothing after it)
            # Escape special regex characters in the title
            title_escaped = re.escape(title)
            title_pattern = title_escaped + r'$'
            title_match = re.search(title_pattern, remaining_text, re.MULTILINE)
            
            if title_match:
                # Start from the line after the title repetition
                start_pos = title_match.end()
                # Skip to next line
                next_line_start = remaining_text.find('\n', start_pos)
                if next_line_start != -1:
                    start_pos = next_line_start + 1
                else:
                    start_pos = len(remaining_text)
                
                # Skip leading whitespace
                while start_pos < len(remaining_text) and remaining_text[start_pos] in ' \n\t':
                    start_pos += 1
                
                body_text_from_start = remaining_text[start_pos:]
            else:
                warnings.append(f"Title repetition not found")
                body_text_from_start = remaining_text
        else:
            warnings.append("Document has only 1 line or is empty")
            body_text_from_start = raw_text
        
        # Find end point: "Author", "Authors", "Footnotes", "Sources", "Related articles", "Learn more here", "Discover more insights here", "Read the full research", "About the author", "Additional authors", or cookie policy text (whichever comes first)
        # Case-insensitive, but must be alone on one line
        author_pattern = r'^Author$'
        authors_pattern = r'^Authors$'
        footnotes_pattern = r'^Footnotes$'
        sources_pattern = r'^Sources$'
        related_articles_pattern = r'^Related articles$'
        learn_more_pattern = r'^Learn more here$'
        discover_more_pattern = r'^Discover more insights here$'
        read_full_research_pattern = r'^Read the full research$'
        about_author_pattern = r'^About the author$'
        additional_authors_pattern = r'^Additional authors$'
        cookie_policy_pattern = r'We and our partners use cookies and other technologies on this site to collect data'
        
        author_match = re.search(author_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        authors_match = re.search(authors_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        footnotes_match = re.search(footnotes_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        sources_match = re.search(sources_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        related_articles_match = re.search(related_articles_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        learn_more_match = re.search(learn_more_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        discover_more_match = re.search(discover_more_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        read_full_research_match = re.search(read_full_research_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        about_author_match = re.search(about_author_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        additional_authors_match = re.search(additional_authors_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        cookie_policy_match = re.search(cookie_policy_pattern, body_text_from_start, re.IGNORECASE)
        
        end_pos = len(body_text_from_start)
        
        # Collect all matches and find the earliest one
        matches = []
        if author_match:
            matches.append(("Author", author_match.start()))
        if authors_match:
            matches.append(("Authors", authors_match.start()))
        if footnotes_match:
            matches.append(("Footnotes", footnotes_match.start()))
        if sources_match:
            matches.append(("Sources", sources_match.start()))
        if related_articles_match:
            matches.append(("Related articles", related_articles_match.start()))
        if learn_more_match:
            matches.append(("Learn more here", learn_more_match.start()))
        if discover_more_match:
            matches.append(("Discover more insights here", discover_more_match.start()))
        if read_full_research_match:
            matches.append(("Read the full research", read_full_research_match.start()))
        if about_author_match:
            matches.append(("About the author", about_author_match.start()))
        if additional_authors_match:
            matches.append(("Additional authors", additional_authors_match.start()))
        if cookie_policy_match:
            matches.append(("Cookie policy text", cookie_policy_match.start()))
        
        if matches:
            # Sort by position and use the first one
            matches.sort(key=lambda x: x[1])
            end_pos = matches[0][1]
        else:
            warnings.append("No end marker found (Author/Authors/Footnotes/Sources/Related articles/Learn more here/Discover more insights here/Read the full research/About the author/Additional authors/Cookie policy)")
        
        # Extract body text
        body_text = body_text_from_start[:end_pos].strip()
        
        # Construct cleaned text: Title + body
        cleaned_text = f"{title}\n\n{body_text}"
        
        # Count words (all words in cleaned text)
        word_count = len(cleaned_text.split())
        
        # Write cleaned text to output file
        output_file_path = os.path.join(OUTPUT_DIR, f"{doc_id}.txt")
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned_text)
        
        # Write to output Excel
        ws_output.cell(row=output_row, column=1, value=doc_id)
        ws_output.cell(row=output_row, column=2, value=firm)
        ws_output.cell(row=output_row, column=3, value=title)
        ws_output.cell(row=output_row, column=4, value="SUCCESS")
        ws_output.cell(row=output_row, column=5, value=" | ".join(warnings) if warnings else "")
        ws_output.cell(row=output_row, column=6, value=word_count)
        output_row += 1
        
        processed += 1
        status_str = f"✓ Processed | Words: {word_count}"
        if warnings:
            status_str += f" | Warnings: {'; '.join(warnings)}"
        print(f"Row {row_idx} (Doc_ID: {doc_id}): {status_str}")
    
    except Exception as e:
        ws_output.cell(row=output_row, column=1, value=doc_id if 'doc_id' in locals() else "")
        ws_output.cell(row=output_row, column=2, value=firm if 'firm' in locals() else "")
        ws_output.cell(row=output_row, column=3, value=title if 'title' in locals() else "")
        ws_output.cell(row=output_row, column=4, value="FAILED")
        ws_output.cell(row=output_row, column=5, value=f"Error during cleanup: {str(e)}")
        output_row += 1
        print(f"Row {row_idx}: Error during cleanup - {str(e)}")
        failed += 1
        continue

# Save the output Excel
wb_output.save(OUTPUT_EXCEL)

print("-" * 80)
print(f"Cleanup complete!")
print(f"Processed: {processed}")
print(f"Skipped (non-RR): {skipped}")
print(f"Failed: {failed}")
print(f"Cleaned files saved to: {OUTPUT_DIR}/")
print(f"Log saved to: {OUTPUT_EXCEL}")

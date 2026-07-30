import openpyxl
import os
from pathlib import Path
import re
from datetime import datetime

# Directories
INPUT_EXCEL = "master.xlsx"
RAW_TEXT_DIR = "txt_files_raw"
OUTPUT_DIR = "txt_files_cleaned_HS"
OUTPUT_EXCEL = "cleanup_HS_log.xlsx"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load the input Excel file (DOCUMENTS tab) - data_only=True to read calculated values, not formulas
wb_input = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
ws_input = wb_input["DOCUMENTS"]

# Create a new Excel file for tracking warnings and word counts
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.title = "Cleanup_Log"

# Set up headers in output Excel
headers = ["Doc_ID", "Firm", "Title", "Status", "Warnings", "Cleaned_Word_Count"]
for col, header in enumerate(headers, start=1):
    ws_output.cell(row=1, column=col, value=header)

# Find column indices in input Excel
firm_col = None
title_col = None
doc_id_col = None

header_row = ws_input[1]
for cell in header_row:
    if cell.value:
        if "Firm" in str(cell.value):
            firm_col = cell.column
        elif "Title" in str(cell.value) or "title" in str(cell.value):
            title_col = cell.column
        elif "Doc_ID" in str(cell.value) or "doc_id" in str(cell.value):
            doc_id_col = cell.column

if not firm_col or not title_col or not doc_id_col:
    print("Error: Could not find required columns (Firm, Title, Doc_ID) in DOCUMENTS tab")
    exit(1)

print(f"Output directory: {OUTPUT_DIR}")
print(f"Output Excel: {OUTPUT_EXCEL}")
print(f"Firm col: {openpyxl.utils.get_column_letter(firm_col)}, Title col: {openpyxl.utils.get_column_letter(title_col)}, Doc_ID col: {openpyxl.utils.get_column_letter(doc_id_col)}")
print("-" * 80)

# Process rows
processed = 0
skipped = 0
failed = 0
output_row = 2

for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=False), start=2):
    try:
        # Get firm, title, and doc_id
        firm_cell = ws_input.cell(row=row_idx, column=firm_col)
        title_cell = ws_input.cell(row=row_idx, column=title_col)
        doc_id_cell = ws_input.cell(row=row_idx, column=doc_id_col)
        
        firm = firm_cell.value
        title = title_cell.value
        doc_id = doc_id_cell.value
        
        # Skip if not HS
        if firm != "HS":
            skipped += 1
            continue
        
        # Skip if missing title or doc_id
        if not title or not doc_id:
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value="Missing title or doc_id")
            output_row += 1
            failed += 1
            continue
        
        # Construct raw text file path (Doc_ID already includes firm prefix like "HS_01")
        raw_file_path = os.path.join(RAW_TEXT_DIR, f"{doc_id}.txt")
        
        # Check if file exists
        if not os.path.exists(raw_file_path):
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value=f"Raw text file not found: {raw_file_path}")
            output_row += 1
            failed += 1
            continue
        
        # Read raw text
        with open(raw_file_path, 'r', encoding='utf-8') as f:
            raw_text = f.read()
        
        # ===== HS-SPECIFIC CLEANUP =====
        warnings = []
        
        # Find start point: 
        # 1. Skip first 3 non-empty lines
        # 2. Search for title repetition (with nothing after it - end of line)
        # 3. Start from line after the title repetition
        
        lines = raw_text.split('\n')
        
        # Skip first 3 non-empty lines
        non_empty_count = 0
        skip_until_idx = 0
        for idx, line in enumerate(lines):
            if line.strip():
                non_empty_count += 1
                if non_empty_count == 3:
                    skip_until_idx = idx + 1
                    break
        
        remaining_text = '\n'.join(lines[skip_until_idx:])
        
        # Search for the title at the end of a line (with nothing after it)
        # Escape special regex characters in the title
        title_escaped = re.escape(title)
        title_pattern = title_escaped + r'$'
        title_match = re.search(title_pattern, remaining_text, re.MULTILINE)
        
        if title_match:
            # Start from the line after the title repetition
            start_pos = title_match.end()
            # Skip to next line
            next_line_start = remaining_text.find('\n', start_pos)
            if next_line_start != -1:
                start_pos = next_line_start + 1
            else:
                start_pos = len(remaining_text)
            
            # Skip leading whitespace
            while start_pos < len(remaining_text) and remaining_text[start_pos] in ' \n\t':
                start_pos += 1
            
            body_text_from_start = remaining_text[start_pos:]
        else:
            warnings.append(f"Title repetition not found")
            body_text_from_start = remaining_text
        
        # Find end point: "View more", "Stay connected", "Subscribe", "About the author(s)", "References", "Related Services & Industries", "Related Content", or "Acknowledgements"
        # Case-insensitive, but must be alone on one line
        view_more_pattern = r'^View more$'
        stay_connected_pattern = r'^Stay connected$'
        subscribe_pattern = r'^Subscribe$'
        about_author_pattern = r'^About the authors?$'  # Matches both "About the author" and "About the authors"
        references_pattern = r'^References$'
        related_services_pattern = r'^Related Services & Industries$'
        related_content_pattern = r'^Related Content$'
        acknowledgements_pattern = r'^Acknowledgements$'
        
        view_match = re.search(view_more_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        stay_match = re.search(stay_connected_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        subscribe_match = re.search(subscribe_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        about_author_match = re.search(about_author_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        references_match = re.search(references_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        related_services_match = re.search(related_services_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        related_content_match = re.search(related_content_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        acknowledgements_match = re.search(acknowledgements_pattern, body_text_from_start, re.MULTILINE | re.IGNORECASE)
        
        end_pos = len(body_text_from_start)
        
        # Collect all matches and find the earliest one
        matches = []
        if view_match:
            matches.append(("View more", view_match.start()))
        if stay_match:
            matches.append(("Stay connected", stay_match.start()))
        if subscribe_match:
            matches.append(("Subscribe", subscribe_match.start()))
        if about_author_match:
            matches.append(("About the author(s)", about_author_match.start()))
        if references_match:
            matches.append(("References", references_match.start()))
        if related_services_match:
            matches.append(("Related Services & Industries", related_services_match.start()))
        if related_content_match:
            matches.append(("Related Content", related_content_match.start()))
        if acknowledgements_match:
            matches.append(("Acknowledgements", acknowledgements_match.start()))
        
        if matches:
            # Sort by position and use the first one
            matches.sort(key=lambda x: x[1])
            end_pos = matches[0][1]
        else:
            warnings.append("No end marker found (View more/Stay connected/Subscribe/About the author(s)/References/Related Services & Industries/Related Content/Acknowledgements)")
        
        # Extract body text
        body_text = body_text_from_start[:end_pos].strip()
        
        # Construct cleaned text: Title + body
        cleaned_text = f"{title}\n\n{body_text}"
        
        # Count words (all words in cleaned text)
        word_count = len(cleaned_text.split())
        
        # Write cleaned text to output file
        output_file_path = os.path.join(OUTPUT_DIR, f"{doc_id}.txt")
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned_text)
        
        # Write to output Excel
        ws_output.cell(row=output_row, column=1, value=doc_id)
        ws_output.cell(row=output_row, column=2, value=firm)
        ws_output.cell(row=output_row, column=3, value=title)
        ws_output.cell(row=output_row, column=4, value="SUCCESS")
        ws_output.cell(row=output_row, column=5, value=" | ".join(warnings) if warnings else "")
        ws_output.cell(row=output_row, column=6, value=word_count)
        output_row += 1
        
        processed += 1
        status_str = f"✓ Processed | Words: {word_count}"
        if warnings:
            status_str += f" | Warnings: {'; '.join(warnings)}"
        print(f"Row {row_idx} (Doc_ID: {doc_id}): {status_str}")
    
    except Exception as e:
        ws_output.cell(row=output_row, column=1, value=doc_id if 'doc_id' in locals() else "")
        ws_output.cell(row=output_row, column=2, value=firm if 'firm' in locals() else "")
        ws_output.cell(row=output_row, column=3, value=title if 'title' in locals() else "")
        ws_output.cell(row=output_row, column=4, value="FAILED")
        ws_output.cell(row=output_row, column=5, value=f"Error during cleanup: {str(e)}")
        output_row += 1
        print(f"Row {row_idx}: Error during cleanup - {str(e)}")
        failed += 1
        continue

# Save the output Excel
wb_output.save(OUTPUT_EXCEL)

print("-" * 80)
print(f"Cleanup complete!")
print(f"Processed: {processed}")
print(f"Skipped (non-HS): {skipped}")
print(f"Failed: {failed}")
print(f"Cleaned files saved to: {OUTPUT_DIR}/")
print(f"Log saved to: {OUTPUT_EXCEL}")

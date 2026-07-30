import openpyxl
import os
from pathlib import Path
import re
from datetime import datetime

# Directories
INPUT_EXCEL = "master.xlsx"
RAW_TEXT_DIR = "txt_files_raw"
OUTPUT_DIR = "txt_files_cleaned_EZ"
OUTPUT_EXCEL = "cleanup_EZ_log.xlsx"

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Load the input Excel file (DOCUMENTS tab) - data_only=True to read calculated values, not formulas
wb_input = openpyxl.load_workbook(INPUT_EXCEL, data_only=True)
ws_input = wb_input["DOCUMENTS"]

# Create a new Excel file for tracking warnings and word counts
wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.title = "Cleanup_Log"

# Set up headers in output Excel
headers = ["Doc_ID", "Firm", "Title", "Status", "Warnings", "Cleaned_Word_Count"]
for col, header in enumerate(headers, start=1):
    ws_output.cell(row=1, column=col, value=header)

# Find column indices in input Excel
firm_col = None
title_col = None
doc_id_col = None

header_row = ws_input[1]
for cell in header_row:
    if cell.value:
        if "Firm" in str(cell.value):
            firm_col = cell.column
        elif "Title" in str(cell.value) or "title" in str(cell.value):
            title_col = cell.column
        elif "Doc_ID" in str(cell.value) or "doc_id" in str(cell.value):
            doc_id_col = cell.column

if not firm_col or not title_col or not doc_id_col:
    print("Error: Could not find required columns (Firm, Title, Doc_ID) in DOCUMENTS tab")
    exit(1)

print(f"Output directory: {OUTPUT_DIR}")
print(f"Output Excel: {OUTPUT_EXCEL}")
print(f"Firm col: {openpyxl.utils.get_column_letter(firm_col)}, Title col: {openpyxl.utils.get_column_letter(title_col)}, Doc_ID col: {openpyxl.utils.get_column_letter(doc_id_col)}")
print("-" * 80)

# Process rows
processed = 0
skipped = 0
failed = 0
output_row = 2

for row_idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=False), start=2):
    try:
        # Get firm, title, and doc_id
        firm_cell = ws_input.cell(row=row_idx, column=firm_col)
        title_cell = ws_input.cell(row=row_idx, column=title_col)
        doc_id_cell = ws_input.cell(row=row_idx, column=doc_id_col)
        
        firm = firm_cell.value
        title = title_cell.value
        doc_id = doc_id_cell.value
        
        # Skip if not EZ
        if firm != "EZ":
            skipped += 1
            continue
        
        # Skip if missing title or doc_id
        if not title or not doc_id:
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value="Missing title or doc_id")
            output_row += 1
            failed += 1
            continue
        
        # Construct raw text file path (Doc_ID already includes firm prefix like "EZ_01")
        raw_file_path = os.path.join(RAW_TEXT_DIR, f"{doc_id}.txt")
        
        # Check if file exists
        if not os.path.exists(raw_file_path):
            ws_output.cell(row=output_row, column=1, value=doc_id)
            ws_output.cell(row=output_row, column=2, value=firm)
            ws_output.cell(row=output_row, column=3, value=title)
            ws_output.cell(row=output_row, column=4, value="FAILED")
            ws_output.cell(row=output_row, column=5, value=f"Raw text file not found: {raw_file_path}")
            output_row += 1
            failed += 1
            continue
        
        # Read raw text
        with open(raw_file_path, 'r', encoding='utf-8') as f:
            raw_text = f.read()
        
        # ===== EZ-SPECIFIC CLEANUP =====
        warnings = []
        
        # Find start point: Try multiple markers in order
        start_pos = 0
        start_marker_found = False
        
        # 1. Try "Share via Email" alone on one line
        share_via_email_pattern = r'^Share via Email$'
        share_match = re.search(share_via_email_pattern, raw_text, re.MULTILINE)
        
        if share_match:
            # Start from first non-empty line after "Share via Email"
            start_pos = share_match.end()
            # Skip to next line
            next_line_start = raw_text.find('\n', start_pos)
            if next_line_start != -1:
                start_pos = next_line_start + 1
            else:
                start_pos = len(raw_text)
            
            # Skip leading whitespace
            while start_pos < len(raw_text) and raw_text[start_pos] in ' \n\t':
                start_pos += 1
            
            start_marker_found = True
        else:
            # 2. Try "X mins read" pattern (any number + "mins read" alone on one line)
            mins_read_pattern = r'^\d+\s+mins?\s+read$'
            mins_match = re.search(mins_read_pattern, raw_text, re.MULTILINE)
            
            if mins_match:
                # Start from first non-empty line after "X mins read"
                start_pos = mins_match.end()
                # Skip to next line
                next_line_start = raw_text.find('\n', start_pos)
                if next_line_start != -1:
                    start_pos = next_line_start + 1
                else:
                    start_pos = len(raw_text)
                
                # Skip leading whitespace
                while start_pos < len(raw_text) and raw_text[start_pos] in ' \n\t':
                    start_pos += 1
                
                start_marker_found = True
            else:
                # 3. Try "日本語" to mark beginning
                japanese_pattern = r'日本語'
                japanese_match = re.search(japanese_pattern, raw_text)
                
                if japanese_match:
                    # Start from first non-empty line after "日本語"
                    start_pos = japanese_match.end()
                    # Skip to next line
                    next_line_start = raw_text.find('\n', start_pos)
                    if next_line_start != -1:
                        start_pos = next_line_start + 1
                    else:
                        start_pos = len(raw_text)
                    
                    # Skip leading whitespace
                    while start_pos < len(raw_text) and raw_text[start_pos] in ' \n\t':
                        start_pos += 1
                    
                    start_marker_found = True
                else:
                    warnings.append("No start marker found ('Share via Email', 'X mins read', or '日本語')")
                    start_pos = 0
        
        # Find end point: Try multiple markers, use whichever comes first
        topics_pattern = r'^Topics Related to this Article$'
        touch_pattern = r'^Get in Touch$'
        subscribe_pattern = r'^Subscribe now$'
        about_pattern = r'^About Egon Zehnder$'
        insights_pattern = r'^Stay up to date with our latest insights\.$'
        
        topics_match = re.search(topics_pattern, raw_text[start_pos:], re.MULTILINE)
        touch_match = re.search(touch_pattern, raw_text[start_pos:], re.MULTILINE)
        subscribe_match = re.search(subscribe_pattern, raw_text[start_pos:], re.MULTILINE)
        about_match = re.search(about_pattern, raw_text[start_pos:], re.MULTILINE)
        insights_match = re.search(insights_pattern, raw_text[start_pos:], re.MULTILINE)
        
        end_pos = len(raw_text)
        end_marker_found = False
        
        # Collect all matches and find the earliest one
        matches = []
        if topics_match:
            matches.append(("Topics Related to this Article", start_pos + topics_match.start()))
        if touch_match:
            matches.append(("Get in Touch", start_pos + touch_match.start()))
        if subscribe_match:
            matches.append(("Subscribe now", start_pos + subscribe_match.start()))
        if about_match:
            matches.append(("About Egon Zehnder", start_pos + about_match.start()))
        if insights_match:
            matches.append(("Stay up to date with our latest insights.", start_pos + insights_match.start()))
        
        if matches:
            # Sort by position and use the first one
            matches.sort(key=lambda x: x[1])
            end_pos = matches[0][1]
            end_marker_found = True
        else:
            warnings.append("No end marker found (Topics/Get in Touch/Subscribe now/About Egon Zehnder/Stay up to date)")
        
        # Extract body text
        body_text = raw_text[start_pos:end_pos].strip()
        
        # Construct cleaned text: Title + body
        cleaned_text = f"{title}\n\n{body_text}"
        
        # Count words (all words in cleaned text)
        word_count = len(cleaned_text.split())
        
        # Write cleaned text to output file
        output_file_path = os.path.join(OUTPUT_DIR, f"{doc_id}.txt")
        with open(output_file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned_text)
        
        # Write to output Excel
        ws_output.cell(row=output_row, column=1, value=doc_id)
        ws_output.cell(row=output_row, column=2, value=firm)
        ws_output.cell(row=output_row, column=3, value=title)
        ws_output.cell(row=output_row, column=4, value="SUCCESS")
        ws_output.cell(row=output_row, column=5, value=" | ".join(warnings) if warnings else "")
        ws_output.cell(row=output_row, column=6, value=word_count)
        output_row += 1
        
        processed += 1
        status_str = f"✓ Processed | Words: {word_count}"
        if warnings:
            status_str += f" | Warnings: {'; '.join(warnings)}"
        print(f"Row {row_idx} (Doc_ID: {doc_id}): {status_str}")
    
    except Exception as e:
        ws_output.cell(row=output_row, column=1, value=doc_id if 'doc_id' in locals() else "")
        ws_output.cell(row=output_row, column=2, value=firm if 'firm' in locals() else "")
        ws_output.cell(row=output_row, column=3, value=title if 'title' in locals() else "")
        ws_output.cell(row=output_row, column=4, value="FAILED")
        ws_output.cell(row=output_row, column=5, value=f"Error during cleanup: {str(e)}")
        output_row += 1
        print(f"Row {row_idx}: Error during cleanup - {str(e)}")
        failed += 1
        continue

# Save the output Excel
wb_output.save(OUTPUT_EXCEL)

print("-" * 80)
print(f"Cleanup complete!")
print(f"Processed: {processed}")
print(f"Skipped (non-EZ): {skipped}")
print(f"Failed: {failed}")
print(f"Cleaned files saved to: {OUTPUT_DIR}/")
print(f"Log saved to: {OUTPUT_EXCEL}")
